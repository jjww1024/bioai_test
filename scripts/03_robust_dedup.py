# -*- coding: utf-8 -*-
"""값 퍼짐(spread) 기반 신뢰도 중복제거 → 시트 'robust_dedup' 추가.
IC50를 pIC50(로그)로 변환해 물질별로 집계하고, 값이 너무 튀는 물질은 폐기 플래그.
"""
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

XL = "data/HSD17B13_IC50_merged.xlsx"

df = pd.read_excel(XL, sheet_name="all_data").dropna(subset=["canonical_smiles", "ic50_nM"])
# 정확한(=) 값만 수치 집계에 사용, 부등호(>,<)는 따로 카운트
rel = df["relation"].astype(str).str.strip()
df["is_exact"] = ~rel.isin([">", "<", ">=", "<="])
df["pIC50"] = 9 - np.log10(df["ic50_nM"].clip(lower=1e-6))


def agg(group):
    ex = group[group["is_exact"]]
    n_exact = len(ex)
    n_qual = len(group) - n_exact
    out = {
        "compound_id": group["compound_id"].dropna().iloc[0] if group["compound_id"].notna().any() else "",
        "smiles": group["smiles"].dropna().iloc[0] if group["smiles"].notna().any() else "",
        "n_exact": n_exact, "n_qualified": n_qual,
        "sources": ", ".join(sorted(group["source"].dropna().unique())),
    }
    if n_exact == 0:
        out.update(ic50_median_nM=np.nan, pIC50_median=np.nan,
                   fold_change=np.nan, range_log=np.nan,
                   confidence="qualified_only", recommend_keep=False)
        return pd.Series(out)
    p = ex["pIC50"]
    med_p = float(p.median())
    rng = float(p.max() - p.min())          # 로그 범위 = log10(max/min)
    fold = 10 ** rng                         # 배수 (max/min)
    if n_exact == 1:
        conf = "single"
    elif rng <= 0.5:
        conf = "high (<=3x)"
    elif rng <= 1.0:
        conf = "good (<=10x)"
    elif rng <= 2.0:
        conf = "moderate (10-100x)"
    else:
        conf = "CONFLICT (>100x)"
    out.update(
        ic50_median_nM=round(10 ** (9 - med_p), 3),
        pIC50_median=round(med_p, 3),
        fold_change=round(fold, 1),
        range_log=round(rng, 2),
        confidence=conf,
        recommend_keep=(conf != "CONFLICT (>100x)"),
    )
    return pd.Series(out)


res = df.groupby("canonical_smiles").apply(agg, include_groups=False).reset_index()
res = res[["canonical_smiles", "compound_id", "smiles", "ic50_median_nM",
           "pIC50_median", "n_exact", "n_qualified", "fold_change", "range_log",
           "confidence", "recommend_keep", "sources"]]
# 튀는 것부터 위로
order = {"CONFLICT (>100x)": 0, "moderate (10-100x)": 1, "good (<=10x)": 2,
         "high (<=3x)": 3, "single": 4, "qualified_only": 5}
res["_o"] = res["confidence"].map(order).fillna(9)
res = res.sort_values(["_o", "fold_change"], ascending=[True, False]).drop(columns="_o").reset_index(drop=True)

# 요약 먼저 출력 (저장 실패해도 결과는 보이게)
print("== robust_dedup 계산 결과 ==")
print("물질 수:", len(res))
print(res["confidence"].value_counts().to_string())
print("\n폐기 권장(>100x 충돌) 예시:")
for _, r in res[res["confidence"] == "CONFLICT (>100x)"].head(5).iterrows():
    print(f"  {r['compound_id']}: {r['fold_change']}배 차이, 측정 {r['n_exact']}건")

# 저장 (원본이 Excel에서 열려 잠겨 있으면 새 파일로)
target = XL
try:
    with pd.ExcelWriter(XL, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        res.to_excel(w, sheet_name="robust_dedup", index=False)
except PermissionError:
    target = XL.replace(".xlsx", "_robust.xlsx")
    print(f"\n[경고] 원본이 열려 있어 저장 불가 → 새 파일로 저장: {target}")
    with pd.ExcelWriter(target, engine="openpyxl") as w:
        res.to_excel(w, sheet_name="robust_dedup", index=False)

# CONFLICT(폐기 권장) 행 빨간색 하이라이트
wb = load_workbook(target)
ws = wb["robust_dedup"]
red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
conf_col = list(res.columns).index("confidence") + 1
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=conf_col).value == "CONFLICT (>100x)":
        for c in range(1, len(res.columns) + 1):
            ws.cell(row=row, column=c).fill = red
wb.save(target)
print("\n저장 완료 →", target)
