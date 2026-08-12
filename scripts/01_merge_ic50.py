# -*- coding: utf-8 -*-
"""BindingDB + ChEMBL TSV에서 SMILES와 IC50 값을 추출·병합하고,
RDKit canonical SMILES로 중복을 판정해 하나의 Excel로 저장한다."""
import re
import numpy as np
import pandas as pd
from rdkit import Chem
from rdkit import RDLogger
RDLogger.DisableLog("rdApp.*")

# 데이터는 프로젝트의 data/ 폴더에 아래 이름으로 두세요 (data/는 .gitignore 처리됨).
BINDINGDB = "data/bindingdb_hsd17b13.tsv"
CHEMBL = "data/chembl_hsd17b13.tsv"
OUT = "data/HSD17B13_IC50_merged.xlsx"


def parse_value(x):
    """'>10000', '<0.5', '100' → (relation, float)."""
    if pd.isna(x):
        return "", np.nan
    s = str(x).strip()
    m = re.match(r"^\s*([<>=~]+)?\s*([0-9.eE+-]+)", s)
    if not m:
        return "", np.nan
    rel = m.group(1) or "="
    try:
        return rel, float(m.group(2))
    except ValueError:
        return rel, np.nan


def canon(smiles):
    """RDKit canonical SMILES. 실패 시 None."""
    if pd.isna(smiles):
        return None
    mol = Chem.MolFromSmiles(str(smiles))
    return Chem.MolToSmiles(mol) if mol else None


# ---------- 1) BindingDB ----------
bdb = pd.read_csv(BINDINGDB, sep="\t", engine="python", on_bad_lines="skip",
                  usecols=["Ligand SMILES", "IC50 (nM)",
                           "BindingDB Ligand Name", "ChEMBL ID of Ligand"])
bdb = bdb[bdb["IC50 (nM)"].notna()].copy()
rel_val = bdb["IC50 (nM)"].apply(parse_value)
bdb["relation"] = [r for r, _ in rel_val]
bdb["ic50_nM"] = [v for _, v in rel_val]
bdb_out = pd.DataFrame({
    "source": "BindingDB",
    "compound_id": bdb["ChEMBL ID of Ligand"].fillna(bdb["BindingDB Ligand Name"]),
    "smiles": bdb["Ligand SMILES"],
    "ic50_nM": bdb["ic50_nM"],
    "relation": bdb["relation"],
    "pChEMBL": np.nan,
    "std_type": "IC50",
})

# ---------- 2) ChEMBL (IC50만) ----------
chembl = pd.read_csv(CHEMBL, sep="\t", engine="python", on_bad_lines="skip",
                     usecols=["Molecule ChEMBL ID", "Smiles", "Standard Type",
                              "Standard Relation", "Standard Value",
                              "Standard Units", "pChEMBL Value"])
chembl = chembl[chembl["Standard Type"].astype(str).str.upper() == "IC50"].copy()
# 단위 nM 통일 (uM 등은 환산)
val = pd.to_numeric(chembl["Standard Value"], errors="coerce")
units = chembl["Standard Units"].astype(str).str.strip()
val_nM = np.where(units == "uM", val * 1000, val)
val_nM = np.where(units == "M", val * 1e9, val_nM)
chembl["ic50_nM"] = np.where(units.isin(["nM", "uM", "M"]), val_nM, np.nan)
chembl_out = pd.DataFrame({
    "source": "ChEMBL",
    "compound_id": chembl["Molecule ChEMBL ID"],
    "smiles": chembl["Smiles"],
    "ic50_nM": chembl["ic50_nM"],
    "relation": chembl["Standard Relation"].astype(str).str.replace("'", "").fillna("="),
    "pChEMBL": pd.to_numeric(chembl["pChEMBL Value"], errors="coerce"),
    "std_type": "IC50",
})

# ---------- 3) 병합 + canonical SMILES ----------
df = pd.concat([bdb_out, chembl_out], ignore_index=True)
df = df[df["smiles"].notna() & (df["smiles"].astype(str).str.len() > 0)].copy()
df["canonical_smiles"] = df["smiles"].apply(canon)
df["valid_structure"] = df["canonical_smiles"].notna()

# ---------- 4) 중복 판정 (canonical 기준) ----------
key = df["canonical_smiles"].fillna(
    pd.Series("INVALID_" + df.index.astype(str), index=df.index))
counts = key.map(key.value_counts())
df["dup_count"] = counts.values
df["is_duplicate"] = df["dup_count"] > 1
# 중복 그룹 번호 (같은 물질끼리 같은 번호)
dup_keys = sorted(k for k, c in key.value_counts().items()
                  if c > 1 and not str(k).startswith("INVALID_"))
gid = {k: i + 1 for i, k in enumerate(dup_keys)}
df["dup_group"] = key.map(gid).astype("Int64")
# 소스 간 중복인지 (BindingDB & ChEMBL 양쪽에 존재)
src_per_key = df.groupby(key)["source"].transform(lambda s: s.nunique())
df["cross_source_dup"] = (src_per_key > 1) & df["is_duplicate"]

# 보기 좋게 정렬: 중복 그룹 먼저, 그 안에서 canonical 묶어서
df = df.sort_values(["is_duplicate", "dup_group", "canonical_smiles", "source"],
                    ascending=[False, True, True, True]).reset_index(drop=True)

col_order = ["source", "compound_id", "smiles", "canonical_smiles", "ic50_nM",
             "relation", "pChEMBL", "std_type", "valid_structure",
             "is_duplicate", "dup_group", "dup_count", "cross_source_dup"]
df = df[col_order]

# ---------- 5) Excel 저장 (전체 + 중복만, 중복행 노란색) ----------
dups = df[df["is_duplicate"]].copy()
with pd.ExcelWriter(OUT, engine="openpyxl") as w:
    df.to_excel(w, sheet_name="all_data", index=False)
    dups.to_excel(w, sheet_name="duplicates_only", index=False)

# 중복 행 노란색 하이라이트
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
wb = load_workbook(OUT)
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ws = wb["all_data"]
dup_col = col_order.index("is_duplicate") + 1
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=dup_col).value is True:
        for c in range(1, len(col_order) + 1):
            ws.cell(row=row, column=c).fill = fill
wb.save(OUT)

# ---------- 요약 ----------
print("저장 완료 →", OUT)
print(f"총 행: {len(df)} (BindingDB {int((df['source']=='BindingDB').sum())}, "
      f"ChEMBL {int((df['source']=='ChEMBL').sum())})")
print(f"파싱 실패(무효 구조): {int((~df['valid_structure']).sum())}")
print(f"중복 물질 그룹 수: {df['dup_group'].dropna().nunique()}")
print(f"중복에 속한 행 수: {int(df['is_duplicate'].sum())}")
print(f"소스 간(BindingDB↔ChEMBL) 중복 행 수: {int(df['cross_source_dup'].sum())}")
print(f"고유 물질 수(canonical 기준): {df['canonical_smiles'].nunique()}")
